import sys, json
sys.stdout.reconfigure(encoding='utf-8')
with open(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\index.html','r',encoding='utf-8') as f:
    d = f.read()

# Extract DATA_PRODUCTS JSON
prod_start = d.find('var DATA_PRODUCTS = [') + len('var DATA_PRODUCTS = [')
prod_end = d.find('];', prod_start)
raw = d[prod_start:prod_end].strip()

# Parse using json
entries = []
depth = 0
start = 0
for i, ch in enumerate(raw):
    if ch == '{':
        if depth == 0: start = i
        depth += 1
    elif ch == '}':
        depth -= 1
        if depth == 0:
            entries.append(raw[start:i+1])

prods = [json.loads(e) for e in entries]

# Check total
print(f'Total products: {len(prods)}')

# Group by code and show standards
codes = {}
for p in prods:
    k = p['code']
    codes.setdefault(k, []).append(p['standard'])

# Check Excel data
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\samples\Gia ban toi thieu.xlsx', data_only=True)
ws = wb['Sheet1']
xl_rows = 0
xl_codes = {}
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if not code: continue
    code = str(code).strip()
    std = str(ws.cell(r, 2).value or '').strip()
    xl_codes[code + '|' + std] = True
    xl_rows += 1

print(f'\nExcel rows: {xl_rows}')
print(f'HTML products: {len(prods)}')

# Check if any missing
in_html = {}
for p in prods:
    in_html[p['code'] + '|' + p['standard']] = True

missing_xl = [k for k in xl_codes if k not in in_html]
extra_html = [k for k in in_html if k not in xl_codes]
if missing_xl:
    print(f'\n⚠️ Missing from HTML: {missing_xl}')
else:
    print('✅ All Excel products present in HTML')

if extra_html:
    print(f'⚠️ Extra in HTML: {extra_html}')
else:
    print('✅ No extra products')

# Show price comparison for first 5 products
print(f'\n--- Sample price comparison (first 5) ---')
for p in prods[:5]:
    print(f'{p["code"]:10s} {p["standard"]:5s} CN_VND={p["exw_vnd_cn"]:>12_} OTHER_VND={p["exw_vnd_other"]:>12_}')
