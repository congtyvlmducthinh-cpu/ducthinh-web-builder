# -*- coding: utf-8 -*-
"""Step 1: Read Excel, build new DATA_PRODUCTS with _cn/_other fields"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\samples\Gia ban toi thieu.xlsx', data_only=True)
ws = wb['Sheet1']

products = []
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if not code: continue
    std = ws.cell(r, 2).value or ''
    machine = ws.cell(r, 3).value or ''
    def v(col):
        val = ws.cell(r, col).value
        return round(val, 2) if val else None
    products.append({
        "code": str(code).strip(), "standard": str(std).strip(), "machine": str(machine).strip(),
        "exw_vnd_cn": v(4), "exw_usd_cn": v(5), "comm_vnd_cn": v(6), "comm_usd_cn": v(7),
        "pkg25_vnd_cn": v(8), "pkg25_usd_cn": v(9), "jumbo_vnd_cn": v(10), "jumbo_usd_cn": v(11),
        "exw_vnd_other": v(12), "exw_usd_other": v(13), "comm_vnd_other": v(14), "comm_usd_other": v(15),
        "pkg25_vnd_other": v(16), "pkg25_usd_other": v(17), "jumbo_vnd_other": v(18), "jumbo_usd_other": v(19),
    })

def fmt(v):
    if v is None: return "null"
    if isinstance(v, float): return str(round(v, 2))
    return json.dumps(str(v))

entries = []
for p in products:
    fields = [
        f'"code":{json.dumps(p["code"])}',
        f'"size":"—"',
        f'"standard":{json.dumps(p["standard"])}',
        f'"machine":{json.dumps(p["machine"])}',
        f'"exw_vnd_cn":{fmt(p["exw_vnd_cn"])}','"exw_usd_cn":{fmt(p["exw_usd_cn"])}',
        f'"comm_vnd_cn":{fmt(p["comm_vnd_cn"])}','"comm_usd_cn":{fmt(p["comm_usd_cn"])}',
        f'"pkg25_vnd_cn":{fmt(p["pkg25_vnd_cn"])}','"pkg25_usd_cn":{fmt(p["pkg25_usd_cn"])}',
        f'"jumbo_vnd_cn":{fmt(p["jumbo_vnd_cn"])}','"jumbo_usd_cn":{fmt(p["jumbo_usd_cn"])}',
        f'"exw_vnd_other":{fmt(p["exw_vnd_other"])}','"exw_usd_other":{fmt(p["exw_usd_other"])}',
        f'"comm_vnd_other":{fmt(p["comm_vnd_other"])}','"comm_usd_other":{fmt(p["comm_usd_other"])}',
        f'"pkg25_vnd_other":{fmt(p["pkg25_vnd_other"])}','"pkg25_usd_other":{fmt(p["pkg25_usd_other"])}',
        f'"jumbo_vnd_other":{fmt(p["jumbo_vnd_other"])}','"jumbo_usd_other":{fmt(p["jumbo_usd_other"])}',
    ]
    entries.append('{' + ','.join(fields) + '}')

output = 'var DATA_PRODUCTS = [' + ','.join(entries) + '];'
with open(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\_new_data_products.txt', 'w', encoding='utf-8') as f:
    f.write(output)
print(f"Done! {len(products)} products written")
print(f"Length: {len(output)} chars")
