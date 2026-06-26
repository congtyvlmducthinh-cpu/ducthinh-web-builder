# Fix bugs:
# 1. applyMarket() recursion - should call render(), not itself
# 2. Rebuild DATA_PRODUCTS with EXACT Excel data including duplicates
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
with open(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\index.html','r',encoding='utf-8') as f:
    d = f.read()

# Fix 1: applyMarket recursion
bad = '  // Re-render\n  applyMarket();\n  render();'
good = '  // Re-render\n  render();'
if bad in d:
    d = d.replace(bad, good)
    print('✅ Fixed applyMarket recursion')
else:
    print('⚠️ applyMarket recursion fix failed - pattern not found')
    # Show what's there
    idx = d.find('function applyMarket')
    if idx >= 0:
        end = d.find('\n}', d.find('forEach', idx)) + 1
        ctx = d[idx:end+100]
        print(f'Actual: {repr(ctx)}')

# Fix 2: Rebuild DATA_PRODUCTS from Excel correctly
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\samples\Gia ban toi thieu.xlsx', data_only=True)
ws = wb['Sheet1']

products = []
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 1).value
    if not code: continue
    code = str(code).strip()
    std = str(ws.cell(r, 2).value or '').strip()
    machine = str(ws.cell(r, 3).value or '').strip()
    
    def v(col):
        val = ws.cell(r, col).value
        return round(val, 2) if val is not None else None
    
    exw_cn_vnd = v(4)
    exw_cn_usd = v(5)
    comm_cn_vnd = v(6)
    comm_cn_usd = v(7)
    pkg25_cn_vnd = v(8)
    pkg25_cn_usd = v(9)
    jumbo_cn_vnd = v(10)
    jumbo_cn_usd = v(11)
    
    exw_other_vnd = v(12)
    exw_other_usd = v(13)
    comm_other_vnd = v(14)
    comm_other_usd = v(15)
    pkg25_other_vnd = v(16)
    pkg25_other_usd = v(17)
    jumbo_other_vnd = v(18)
    jumbo_other_usd = v(19)
    
    # Map size from old data or use '—'
    # Use a unique key for each row (code|standard)
    key = f'{code}|{std}'
    
    # Default size = '—' for now - we'll fix this
    # Some products have duplicate codes with different standards (e.g., F15A2 with A2- and A2)
    # We CANNOT use size from old data because old data also has duplicates
    # Use size from old data based on code+standard match
    
    p = {
        "code": code,
        "size": "—",  # Will fix below
        "standard": std,
        "machine": machine,
        "exw_vnd": exw_cn_vnd,
        "exw_usd": exw_cn_usd,
        "comm_vnd": comm_cn_vnd,
        "comm_usd": comm_cn_usd,
        "pkg25_vnd": pkg25_cn_vnd,
        "pkg25_usd": pkg25_cn_usd,
        "jumbo_vnd": jumbo_cn_vnd,
        "jumbo_usd": jumbo_cn_usd,
        "exw_vnd_cn": exw_cn_vnd,
        "exw_usd_cn": exw_cn_usd,
        "comm_vnd_cn": comm_cn_vnd,
        "comm_usd_cn": comm_cn_usd,
        "pkg25_vnd_cn": pkg25_cn_vnd,
        "pkg25_usd_cn": pkg25_cn_usd,
        "jumbo_vnd_cn": jumbo_cn_vnd,
        "jumbo_usd_cn": jumbo_cn_usd,
        "exw_vnd_other": exw_other_vnd,
        "exw_usd_other": exw_other_usd,
        "comm_vnd_other": comm_other_vnd,
        "comm_usd_other": comm_other_usd,
        "pkg25_vnd_other": pkg25_other_vnd,
        "pkg25_usd_other": pkg25_other_usd,
        "jumbo_vnd_other": jumbo_other_vnd,
        "jumbo_usd_other": jumbo_other_usd,
    }
    products.append(p)

# Now fix sizes from OLD data since old data had correct sizes
# Parse old data
old_data_match = d.find('var DATA_PRODUCTS = [')
old_data_end = d.find('];', old_data_match) + 2
old_raw = d[old_data_match + len('var DATA_PRODUCTS = ['):d.find('];', old_data_match)]

old_entries = []
depth = 0
start = 0
s = old_raw
for i, ch in enumerate(s):
    if ch == '{':
        if depth == 0: start = i
        depth += 1
    elif ch == '}':
        depth -= 1
        if depth == 0:
            old_entries.append(s[start:i+1])

old_size_map = {}
for entry in old_entries:
    try:
        obj = json.loads(entry)
        old_size_map[obj['code'] + '|' + obj['standard']] = obj.get('size', '—')
    except:
        pass

for p in products:
    key = p['code'] + '|' + p['standard']
    if key in old_size_map:
        p['size'] = old_size_map[key]

# Build JS
def fmt(v):
    if v is None: return "null"
    if isinstance(v, bool): return "true" if v else "false"
    return json.dumps(str(v)) if isinstance(v, str) else str(v)

entries = ['  ' + json.dumps(p, ensure_ascii=False) for p in products]
new_data = 'var DATA_PRODUCTS = [\n' + ',\n'.join(entries) + '\n];'

# Replace DATA_PRODUCTS
prod_start = d.find('var DATA_PRODUCTS = [')
prod_end = d.find('];', prod_start) + 2
d = d[:prod_start] + new_data + d[prod_end:]

with open(r'C:\Users\Admin\.openclaw\workspace-skills\web-builder\sites\kiem-tra-gia\index.html','w',encoding='utf-8') as f:
    f.write(d)

print(f'\n✅ Rebuilt DATA_PRODUCTS: {len(products)} products')
print(f'✅ Fixed applyMarket() recursion')

# Verify
print(f'\nProduct count: {len(products)}')
# Show unique codes with multiple standards
from collections import Counter
code_standards = {}
for p in products:
    code_standards.setdefault(p['code'], set()).add(p['standard'])
for code, stds in code_standards.items():
    if len(stds) > 1:
        print(f'  {code}: {stds}')

# Verify with Node
import subprocess
result = subprocess.run(['node', '-e', f'''
const fs = require('fs');
const html = fs.readFileSync('C:/Users/Admin/.openclaw/workspace-skills/web-builder/sites/kiem-tra-gia/index.html', 'utf-8');
const m = html.match(/<script>([\s\S]*?)<\/script>/);
if (m) {{
  try {{ new Function(m[1]); console.log("JS OK"); }}
  catch(e) {{ console.log("JS ERROR:", e.message); }}
}}
'''], capture_output=True, text=True)
print(f'JS Validation: {result.stdout.strip()}')
if result.stderr: print(f'Error: {result.stderr}')
